"""
Report generators for different formats (PDF, Excel, CSV)
"""
import os
import csv
from datetime import datetime
from io import BytesIO
from django.conf import settings
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.pdfgen import canvas
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from django.utils import timezone


class BasePDFGenerator:
    """Base class for PDF report generation"""
    
    def __init__(self, report):
        self.report = report
        self.user = report.user
        self.vehicle = report.vehicle
        self.styles = getSampleStyleSheet()
        self.elements = []
        
        # Custom styles
        self.title_style = ParagraphStyle(
            'CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1f2937'),
            spaceAfter=30,
        )
        
        self.heading_style = ParagraphStyle(
            'CustomHeading',
            parent=self.styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#3b82f6'),
            spaceAfter=12,
        )
    
    def generate(self, output_path):
        """Generate PDF report"""
        doc = SimpleDocTemplate(
            output_path,
            pagesize=A4,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=18,
        )
        
        # Add header
        self.add_header()
        
        # Add content (implemented by subclasses)
        self.add_content()
        
        # Build PDF
        doc.build(self.elements)
        
        return output_path
    
    def add_header(self):
        """Add report header"""
        # Title
        title = f"{self.report.get_report_type_display()}"
        self.elements.append(Paragraph(title, self.title_style))
        
        # Report info
        info_data = [
            ['Generated:', datetime.now().strftime('%Y-%m-%d %H:%M')],
            ['User:', self.user.get_full_name() or self.user.username],
        ]
        
        if self.vehicle:
            info_data.append([
                'Vehicle:',
                f"{self.vehicle.make} {self.vehicle.model} ({self.vehicle.license_plate})"
            ])
        
        if self.report.date_from or self.report.date_to:
            period = f"{self.report.date_from or 'Start'} to {self.report.date_to or 'End'}"
            info_data.append(['Period:', period])
        
        info_table = Table(info_data, colWidths=[1.5*inch, 4*inch])
        info_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.grey),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ]))
        
        self.elements.append(info_table)
        self.elements.append(Spacer(1, 20))
    
    def add_content(self):
        """Add report content - to be implemented by subclasses"""
        pass
    
    def add_table(self, title, data, col_widths=None):
        """Helper to add a styled table"""
        if title:
            self.elements.append(Paragraph(title, self.heading_style))
        
        table = Table(data, colWidths=col_widths)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
        ]))
        
        self.elements.append(table)
        self.elements.append(Spacer(1, 20))


class VehicleSummaryPDFGenerator(BasePDFGenerator):
    """Generate vehicle summary PDF report"""
    
    def add_content(self):
        from maintenances.models import Maintenance
        from diagnostics.models import Diagnostic
        
        if not self.vehicle:
            self.elements.append(Paragraph("No vehicle specified", self.styles['Normal']))
            return
        
        # Vehicle details
        self.elements.append(Paragraph("Vehicle Information", self.heading_style))
        vehicle_data = [
            ['Make', 'Model', 'Year', 'License Plate', 'VIN'],
            [
                self.vehicle.make,
                self.vehicle.model,
                str(self.vehicle.year),
                self.vehicle.license_plate,
                self.vehicle.vin or 'N/A'
            ]
        ]
        self.add_table(None, vehicle_data)
        
        # Maintenance summary
        maintenances = Maintenance.objects.filter(vehicle=self.vehicle)
        if self.report.date_from:
            maintenances = maintenances.filter(date__gte=self.report.date_from)
        if self.report.date_to:
            maintenances = maintenances.filter(date__lte=self.report.date_to)
        
        total_cost = sum(m.cost for m in maintenances if m.cost)
        
        self.elements.append(Paragraph("Maintenance Summary", self.heading_style))
        summary_data = [
            ['Total Maintenances', 'Total Cost', 'Average Cost'],
            [
                str(maintenances.count()),
                f"€{total_cost:.2f}",
                f"€{total_cost / maintenances.count():.2f}" if maintenances.count() > 0 else '€0.00'
            ]
        ]
        self.add_table(None, summary_data)
        
        # Recent maintenances
        if self.report.include_details and maintenances.exists():
            self.elements.append(Paragraph("Recent Maintenances", self.heading_style))
            maint_data = [['Date', 'Type', 'Description', 'Cost']]
            for maint in maintenances[:10]:
                maint_data.append([
                    maint.date.strftime('%Y-%m-%d'),
                    maint.maintenance_type,
                    maint.description[:50] + '...' if len(maint.description) > 50 else maint.description,
                    f"€{maint.cost:.2f}" if maint.cost else 'N/A'
                ])
            self.add_table(None, maint_data, col_widths=[1*inch, 1.5*inch, 2.5*inch, 1*inch])


class BaseExcelGenerator:
    """Base class for Excel report generation"""
    
    def __init__(self, report):
        self.report = report
        self.user = report.user
        self.vehicle = report.vehicle
        self.wb = Workbook()
        self.ws = self.wb.active
        
        # Styles
        self.header_font = Font(bold=True, size=14, color="FFFFFF")
        self.header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        self.title_font = Font(bold=True, size=18)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def generate(self, output_path):
        """Generate Excel report"""
        # Add header
        self.add_header()
        
        # Add content (implemented by subclasses)
        self.add_content()
        
        # Save workbook
        self.wb.save(output_path)
        
        return output_path
    
    def add_header(self):
        """Add report header"""
        self.ws.title = "Report"
        
        # Title
        self.ws['A1'] = self.report.get_report_type_display()
        self.ws['A1'].font = self.title_font
        
        # Report info
        self.ws['A3'] = 'Generated:'
        self.ws['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M')
        
        self.ws['A4'] = 'User:'
        self.ws['B4'] = self.user.get_full_name() or self.user.username
        
        row = 5
        if self.vehicle:
            self.ws[f'A{row}'] = 'Vehicle:'
            self.ws[f'B{row}'] = f"{self.vehicle.make} {self.vehicle.model} ({self.vehicle.license_plate})"
            row += 1
        
        if self.report.date_from or self.report.date_to:
            self.ws[f'A{row}'] = 'Period:'
            self.ws[f'B{row}'] = f"{self.report.date_from or 'Start'} to {self.report.date_to or 'End'}"
            row += 1
        
        self.current_row = row + 2
    
    def add_content(self):
        """Add report content - to be implemented by subclasses"""
        pass
    
    def add_table_header(self, headers):
        """Add styled table header"""
        for col, header in enumerate(headers, start=1):
            cell = self.ws.cell(row=self.current_row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.border
        
        self.current_row += 1


class VehicleSummaryExcelGenerator(BaseExcelGenerator):
    """Generate vehicle summary Excel report"""
    
    def add_content(self):
        from maintenances.models import Maintenance
        
        if not self.vehicle:
            self.ws.cell(row=self.current_row, column=1, value="No vehicle specified")
            return
        
        # Vehicle details
        self.ws.cell(row=self.current_row, column=1, value="Vehicle Information").font = Font(bold=True, size=12)
        self.current_row += 1
        
        vehicle_headers = ['Make', 'Model', 'Year', 'License Plate', 'VIN']
        self.add_table_header(vehicle_headers)
        
        vehicle_data = [
            self.vehicle.make,
            self.vehicle.model,
            self.vehicle.year,
            self.vehicle.license_plate,
            self.vehicle.vin or 'N/A'
        ]
        
        for col, value in enumerate(vehicle_data, start=1):
            cell = self.ws.cell(row=self.current_row, column=col, value=value)
            cell.border = self.border
        
        self.current_row += 3
        
        # Maintenances
        maintenances = Maintenance.objects.filter(vehicle=self.vehicle)
        if self.report.date_from:
            maintenances = maintenances.filter(date__gte=self.report.date_from)
        if self.report.date_to:
            maintenances = maintenances.filter(date__lte=self.report.date_to)
        
        self.ws.cell(row=self.current_row, column=1, value="Maintenance History").font = Font(bold=True, size=12)
        self.current_row += 1
        
        maint_headers = ['Date', 'Type', 'Description', 'Mileage', 'Cost']
        self.add_table_header(maint_headers)
        
        for maint in maintenances:
            row_data = [
                maint.date.strftime('%Y-%m-%d'),
                maint.maintenance_type,
                maint.description,
                maint.mileage or 'N/A',
                maint.cost or 0
            ]
            
            for col, value in enumerate(row_data, start=1):
                cell = self.ws.cell(row=self.current_row, column=col, value=value)
                cell.border = self.border
            
            self.current_row += 1


class CSVGenerator:
    """Generate CSV reports"""
    
    def __init__(self, report):
        self.report = report
        self.user = report.user
        self.vehicle = report.vehicle
    
    def generate(self, output_path):
        """Generate CSV report"""
        with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            
            # Header
            writer.writerow([self.report.get_report_type_display()])
            writer.writerow(['Generated:', datetime.now().strftime('%Y-%m-%d %H:%M')])
            writer.writerow(['User:', self.user.get_full_name() or self.user.username])
            
            if self.vehicle:
                writer.writerow([
                    'Vehicle:',
                    f"{self.vehicle.make} {self.vehicle.model} ({self.vehicle.license_plate})"
                ])
            
            writer.writerow([])  # Empty row
            
            # Content based on report type
            self.add_content(writer)
        
        return output_path
    
    def add_content(self, writer):
        """Add CSV content based on report type"""
        from maintenances.models import Maintenance
        from diagnostics.models import Diagnostic
        
        if self.report.report_type == 'vehicle_summary':
            self.add_vehicle_summary(writer)
        elif self.report.report_type == 'maintenance_history':
            self.add_maintenance_history(writer)
        elif self.report.report_type == 'diagnostic_history':
            self.add_diagnostic_history(writer)
    
    def add_vehicle_summary(self, writer):
        """Add vehicle summary to CSV"""
        if not self.vehicle:
            writer.writerow(['No vehicle specified'])
            return
        
        writer.writerow(['Vehicle Information'])
        writer.writerow(['Make', 'Model', 'Year', 'License Plate', 'VIN'])
        writer.writerow([
            self.vehicle.make,
            self.vehicle.model,
            self.vehicle.year,
            self.vehicle.license_plate,
            self.vehicle.vin or 'N/A'
        ])
        writer.writerow([])
        
        # Maintenances
        from maintenances.models import Maintenance
        maintenances = Maintenance.objects.filter(vehicle=self.vehicle)
        
        writer.writerow(['Maintenance History'])
        writer.writerow(['Date', 'Type', 'Description', 'Mileage', 'Cost'])
        
        for maint in maintenances:
            writer.writerow([
                maint.date.strftime('%Y-%m-%d'),
                maint.maintenance_type,
                maint.description,
                maint.mileage or 'N/A',
                maint.cost or 0
            ])
    
    def add_maintenance_history(self, writer):
        """Add maintenance history to CSV"""
        from maintenances.models import Maintenance
        
        maintenances = Maintenance.objects.filter(user=self.user)
        
        if self.vehicle:
            maintenances = maintenances.filter(vehicle=self.vehicle)
        if self.report.date_from:
            maintenances = maintenances.filter(date__gte=self.report.date_from)
        if self.report.date_to:
            maintenances = maintenances.filter(date__lte=self.report.date_to)
        
        writer.writerow(['Maintenance History'])
        writer.writerow(['Vehicle', 'Date', 'Type', 'Description', 'Mileage', 'Cost', 'Status'])
        
        for maint in maintenances:
            writer.writerow([
                f"{maint.vehicle.make} {maint.vehicle.model}",
                maint.date.strftime('%Y-%m-%d'),
                maint.maintenance_type,
                maint.description,
                maint.mileage or 'N/A',
                maint.cost or 0,
                maint.status
            ])
    
    def add_diagnostic_history(self, writer):
        """Add diagnostic history to CSV"""
        from diagnostics.models import Diagnostic
        
        diagnostics = Diagnostic.objects.filter(user=self.user)
        
        if self.vehicle:
            diagnostics = diagnostics.filter(vehicle=self.vehicle)
        if self.report.date_from:
            diagnostics = diagnostics.filter(date__gte=self.report.date_from)
        if self.report.date_to:
            diagnostics = diagnostics.filter(date__lte=self.report.date_to)
        
        writer.writerow(['Diagnostic History'])
        writer.writerow(['Vehicle', 'Date', 'Code', 'Description', 'Severity', 'Status'])
        
        for diag in diagnostics:
            writer.writerow([
                f"{diag.vehicle.make} {diag.vehicle.model}",
                diag.date.strftime('%Y-%m-%d'),
                diag.diagnostic_code,
                diag.description,
                diag.severity,
                'Resolved' if diag.resolved else 'Unresolved'
            ])
